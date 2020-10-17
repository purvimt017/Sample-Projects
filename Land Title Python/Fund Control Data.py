# -*- coding: utf-8 -*-
"""
Created on Wed Mar 18 13:21:34 2020

@author: mpurvis
"""
import smtplib, ssl
import config1
import smtplib, ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import config1

def sendemailnotificationfailure():
    sender_email = config1.username
    receiver_email = config1.username
    password = config1.password2

    message = MIMEMultipart("alternative")
    message["Subject"] = "Failed Scheduled Refresh for FundControl"
    message["From"] = sender_email
    message["To"] = receiver_email

    # Create the plain-text and HTML version of your message
    text = "Your scheduled refresh for FundControl has failed."

    # Turn these into plain/html MIMEText objects
    part1 = MIMEText(text, "plain")

    # Add HTML/plain-text parts to MIMEMultipart message
    # The email client will try to render the last part first
    message.attach(part1)

    # Create secure connection with server and send email
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(sender_email, password)
        server.sendmail(
            sender_email, receiver_email, message.as_string()
        )

try:
    import requests
    import pandas as pd
    from requests.auth import HTTPBasicAuth
    from datetime import date
    from openpyxl import load_workbook
    from openpyxl import Workbook
    import sharepy
    import config1
    import os

    def sendemailnotificationsuccess():
        sender_email = config1.username
        receiver_email = config1.username
        password = config1.password2

        message = MIMEMultipart("alternative")
        message["Subject"] = "Successful Scheduled Refresh for FundControl"
        message["From"] = sender_email
        message["To"] = receiver_email

        # Create the plain-text and HTML version of your message
        text = "Your scheduled refresh for FundControl has succeeded."

        # Turn these into plain/html MIMEText objects
        part1 = MIMEText(text, "plain")

        # Add HTML/plain-text parts to MIMEMultipart message
        # The email client will try to render the last part first
        message.attach(part1)

        # Create secure connection with server and send email
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
            server.login(sender_email, password)
            server.sendmail(
                sender_email, receiver_email, message.as_string()
            )

    def getcodeinfo(code,df):
        data = df.query('Code==@code').reset_index()
        DocName = data.iloc[0,2]
        DocType = data.iloc[0,3]
        DocDesc = data.iloc[0,4]
        return DocName, DocType, DocDesc

    def getborrower(df):
        try:
            var = 'Borrower'
            data = playersdf.query('player_type_name==@var').reset_index()
            company_name = data.loc[0,'company_name']
            contact_name = data.loc[0, 'contact_name']
            if(company_name ==''):
                name = contact_name
            else:
                name = company_name
        except KeyError:
            name = ''
        return name

    def getlender(df):
        try:
            var = 'Lender'
            data = playersdf.query('player_type_name==@var').reset_index()
            company_name = data.loc[0,'company_name']
            contact_name = data.loc[0, 'contact_name']
            if(company_name ==''):
                name = contact_name
            else:
                name = company_name
        except KeyError:
            name = ''
        return name

    def getcontractor(df):
        try:
            var = 'Prime Contractor'
            data = playersdf.query('player_type_name==@var').reset_index()
            company_name = data.loc[0,'company_name']
            contact_name = data.loc[0, 'contact_name']
            if(company_name ==''):
                name = contact_name
            else:
                name = company_name
        except KeyError:
            name = ''
        return name

    def getinspector(df):
        try:
            var = 'Inspector'
            data = playersdf.query('player_type_name==@var').reset_index()
            company_name = data.loc[0,'company_name']
            contact_name = data.loc[0, 'contact_name']
            if(company_name ==''):
                name = contact_name
            else:
                name = company_name
        except KeyError:
            name = ''
        return name
    
    def connectsp():
        s = sharepy.connect("https://ltsystems.sharepoint.com", username = config1.username, password = config1.password)
        return s
    
    def downloadlatestwb():
        t = sharepy.connect("https://ltsystems.sharepoint.com", username = config1.username, password = config1.password)
        t.getfile('https://ltsystems.sharepoint.com/sites/PowerBIAdmin45/Shared%20Documents/Source_Data_Files/Fund Control Document Type Codes.xlsx'
                  ,filename="Fund Control Document Type Codes.xlsx")
    def postwb(s):
        with open('Fund Control Field Data.xlsx', 'rb') as file_input:
            try:
                response = s.post(
                    url='https://ltsystems.sharepoint.com' + "/sites/" + "PowerBIAdmin45" 
                    + "/_api/web/GetFolderByServerRelativeUrl('" + 'Shared%20Documents/Source_Data_Files' + "')/Files/add(url='"
                    + 'Fund Control Field Data.xlsx' + "',overwrite=true)",
                    data=file_input)
                status_code = response.status_code
                if response.status_code == 200:
                    print(str(status_code) + ': Post Successful')
            except Exception as err:
                print("Some error occurred: " + str(err))
        with open('Fund Control Financials.xlsx', 'rb') as file_input:
            try:
                response = s.post(
                    url='https://ltsystems.sharepoint.com' + "/sites/" + "PowerBIAdmin45" 
                    + "/_api/web/GetFolderByServerRelativeUrl('" 
                    + 'Shared%20Documents/Source_Data_Files' + "')/Files/add(url='"
                    + 'Fund Control Financials.xlsx' + "',overwrite=true)",
                    data=file_input)
                status_code = response.status_code
                if response.status_code == 200:
                    print(str(status_code) + ': Post Successful')
            except Exception as err:
                print("Some error occurred: " + str(err))
                
    def deletewb():
        if os.path.exists("Fund Control Field Data.xlsx"):
          os.remove("Fund Control Field Data.xlsx")
        else:
          print("The file does not exist")
        if os.path.exists("Fund Control Financials.xlsx"):
          os.remove("Fund Control Financials.xlsx")
        else:
          print("The file does not exist")
        if os.path.exists("Fund Control Document Type Codes.xlsx"):
          os.remove("Fund Control Document Type Codes.xlsx")
        else:
          print("The file does not exist")
          
    downloadlatestwb()
    codes = pd.read_excel("Fund Control Document Type Codes.xlsx")

    #Get Token
    Base_URL = "https://ltgc.fundcontrol.net/fmi/data/v1/databases/fc_data/sessions"

    #Headers
    PARAMS = {'Content-Type':'application/json'}

    #Basic Auth Request
    r = requests.post(url = Base_URL, auth=HTTPBasicAuth(config1.ltuser, config1.ltpass), headers = PARAMS)

    #Extracting Token
    data = r.json()
    tokendic =  data['response']
    token = tokendic['token']

    #getting data
    url = "https://ltgc.fundcontrol.net/fmi/data/v1/databases/fc_data/layouts/api_loan_find/_find"
    headers = {'Content-Type':'application/json', 'Authorization':'Bearer {}'.format(token)}
    param = {'query':[{'loan_number':'*'}], 'limit':10000,
    "limit.Financial_Loans_lns": "1000"}
    r = requests.post(url = url, headers = headers, json = param)
    data2 = r.json()

    response = data2['response']
    respdata = response['data']
    length = len(respdata)

    active = []
    borrower = []
    lender = []
    inspector = []
    contractor = []
    contract_type = []
    date_appraised = []
    date_end_of_draw_period = []
    date_closed = []
    date_inspection_last_entered = []
    date_loan_final = []
    date_matured = []
    department_number = []
    inspection_frequency = []
    loan_interest_when_to_bill = []
    loan_name = []
    loan_number = []
    loan_number_system = []
    loan_status = []
    loan_type = []
    project_description = []
    property_address= []
    property_address_city= []
    property_address_state= []
    property_address_zip= []
    property_type= []
    revolving= []
    text_notes= []

    loan_financials = []
    recordId= []
    chk_dpst_account= []
    chk_dpst_date= []
    chk_dpst_date_cleared= []
    chk_dpst_number= []
    chk_dpst_number_display= []
    chk_dpst_value= []
    doc_co_num= []
    doc_code= []
    doc_date= []
    doc_value= []
    id= []
    vendor_code= []
    vendor_name= []
    vendorjoint_code= []
    vendorjoint_name_address= []
    what_to_do= []
    z_co_code= []
    z_document_code= []
    modId= []
    DocNames = []
    DocTypes = []
    DocDescs = []

    index = 0
    for i in range(length):
        respdatai = respdata[index]
        fieldData = respdatai['fieldData']
        portalData = respdatai['portalData']
        players = portalData['Players']
        playersdf = pd.DataFrame.from_dict(players)
        playersdf.columns = ['recordId', 'active', 'address_singleline',
           'budget_code', 'company_name',
           'contact_code', 'contact_code_related',
           'contact_name', 'contact_name_initials',
           'id', 'loan_number', 'loan_number_system',
           'notes', 'player_status', 'player_type_name',
           'PlayerAcctNum', 'PlayerAgreeDt',
           'PlayerBankNum', 'PlayerIntTyp',
           'PlayerNote', 'PlayerPartStartDt',
           'PlayerPaybyEFT', 'PlayerPmtFrq',
           'PlayerPosition', 'primary_flag',
           'primary_secondary_flag', 'sublot',
           'their_code_for_this_project', 'type_number',
           'contact_email', 'contact_tin', 'modId']
        active.append(fieldData['active'])
        borrower.append(getborrower(playersdf))
        contractor.append(getcontractor(playersdf))
        contract_type.append(fieldData['contract_type'])
        date_appraised.append(fieldData['date_appraised'])
        date_closed.append(fieldData['date_closed'])
        date_end_of_draw_period.append(fieldData['date_end_of_draw_period'])
        date_inspection_last_entered.append(fieldData['date_inspection_last_entered'])
        date_loan_final.append(fieldData['date_loan_final'])
        date_matured.append(fieldData['date_matured'])
        department_number.append(fieldData['department_number'])
        inspector.append(getinspector(playersdf))
        inspection_frequency.append(fieldData['inspection_frequency'])
        lender.append(getlender(playersdf))
        loan_interest_when_to_bill.append(fieldData['loan_interest_when_to_bill'])
        loan_name.append(fieldData['loan_name'])
        loan_num = fieldData['loan_number']
        loan_number.append(loan_num)
        loan_number_system.append(fieldData['loan_number_system'])
        loan_status.append(fieldData['loan_status'])
        loan_type.append(fieldData['loan_type'])
        project_description.append(fieldData['project_description'])
        property_address.append(fieldData['property_address'])
        property_address_city.append(fieldData['property_address_city'])
        property_address_state.append(fieldData['property_address_state'])
        property_address_zip.append(fieldData['property_address_zip'])
        property_type.append(fieldData['property_type'])
        revolving.append(fieldData['revolving'])
        text_notes.append(fieldData['text_notes'])
        index += 1


    fieldDatadf = pd.DataFrame({'active':active,
                              'borrower':borrower,
                              'contractor':contractor,
                              'contract_type':contract_type,
                              'date_appraised':date_appraised,
                              'date_closed':date_closed,
                              'date_end_of_draw_period':date_end_of_draw_period,
                              'date_inspection_last_entered':date_inspection_last_entered,
                              'date_loan_final':date_loan_final,
                              'date_matured':date_matured,
                              'department_number':department_number,
                              'inspector':inspector,
                              'inspection_frequency':inspection_frequency,
                              'lender':lender,
                              'loan_interest_when_to_bill':loan_interest_when_to_bill,
                              'loan_name':loan_name,
                              'loan_number':loan_number,
                              'loan_number_system':loan_number_system,
                              'loan_status':loan_status,
                              'loan_type':loan_type,
                              'project_description':project_description,
                              'property_address':property_address,
                              'property_address_city':property_address_city,
                              'property_address_state':property_address_state,
                              'property_address_zip':property_address_zip,
                              'property_type':property_type,
                              'revolving':revolving,
                              'text_notes':text_notes
                              } )
    '''
    loannum = '70660216CL'
    fieldDatadf = fieldDatadf.query('loan_number==@loannum')
    '''


    index1 = 0
    for i in range(length):
        respdatai = respdata[index1]
        fieldData = respdatai['fieldData']
        loan_num = fieldData['loan_number']
        portalData = respdatai['portalData']
        financial_Loans_lns_list = portalData['Financial_Loans_lns']
        if(len(financial_Loans_lns_list) > 0):
            financial_Loans_lns_df = pd.DataFrame.from_dict(financial_Loans_lns_list)
        else:
            financial_Loans_lns_df = pd.DataFrame(columns=
                    ['recordId', 'chk_dpst_account',
               'chk_dpst_date',
               'chk_dpst_date_cleared',
               'chk_dpst_number',
               'chk_dpst_number_display',
               'chk_dpst_value',
               'doc_co_num', 'doc_code',
               'doc_date', 'doc_value',
               'id', 'vendor_code',
               'vendor_name',
               'vendorjoint_code',
               'vendorjoint_name_address',
               'what_to_do', 'z_co_code',
               'z_document_code', 'modId']
                    )
            financial_Loans_lns_df.append({'recordId': '', 'chk_dpst_account':'',
               'chk_dpst_date':'',
               'chk_dpst_date_cleared':'',
               'chk_dpst_number':'',
               'chk_dpst_number_display':'',
               'chk_dpst_value':'',
               'doc_co_num':'', 'doc_code':'',
               'doc_date':'', 'doc_value':'',
               'id':'', 'vendor_code':'',
               'vendor_name':'',
               'vendorjoint_code':'',
               'vendorjoint_name_address':'',
               'what_to_do':'', 'z_co_code':'',
               'z_document_code':'', 'modId':''}, ignore_index = True)
        financial_Loans_lns_df.columns = ['recordId', 'chk_dpst_account',
           'chk_dpst_date',
           'chk_dpst_date_cleared',
           'chk_dpst_number',
           'chk_dpst_number_display',
           'chk_dpst_value',
           'doc_co_num', 'doc_code',
           'doc_date', 'doc_value',
           'id', 'vendor_code',
           'vendor_name',
           'vendorjoint_code',
           'vendorjoint_name_address',
           'what_to_do', 'z_co_code',
           'z_document_code', 'modId']
        for index, row in financial_Loans_lns_df.iterrows():
            loan_financials.append(loan_num)
            recordId.append(row['recordId'])
            chk_dpst_account.append(row['chk_dpst_account'])
            chk_dpst_date.append(row['chk_dpst_date'])
            chk_dpst_date_cleared.append(row['chk_dpst_date_cleared'])
            chk_dpst_number.append(row['chk_dpst_number'])
            chk_dpst_number_display.append(row['chk_dpst_number_display'])
            chk_dpst_value.append(row['chk_dpst_value'])
            doc_co_num.append(row['doc_co_num'])
            doc_code.append(row['doc_code'])
            doc_date.append(row['doc_date'])
            doc_value.append(row['doc_value'])
            id.append(row['id'])
            vendor_code.append(row['vendor_code'])
            vendor_name.append(row['vendor_name'])
            vendorjoint_code.append(row['vendorjoint_code'])
            vendorjoint_name_address.append(row['vendorjoint_name_address'])
            what_to_do.append(row['what_to_do'])
            z_co_code.append(row['z_co_code'])
            z_document_code.append(row['z_document_code'])
            modId.append(row['modId'])
            code = row['z_document_code']
            try:
                DocName, DocType, DocDesc = getcodeinfo(code, codes)
            except IndexError:
                DocName = ''
                DocType = ''
                DocDesc = ''
            DocNames.append(DocName)
            DocTypes.append(DocType)
            DocDescs.append(DocDesc)
        index1 += 1

    financial_Loans_lns = pd.DataFrame({'loan_number': loan_financials,
                                        'recordId':recordId,
                                        'chk_dpst_account':chk_dpst_account,
                                        'chk_dpst_date':chk_dpst_date,
                                        'chk_dpst_date_cleared':chk_dpst_date_cleared,
                                        'chk_dpst_number':chk_dpst_number,
                                        'chk_dpst_number_display':chk_dpst_number_display,
                                        'chk_dpst_value':chk_dpst_value,
                                        'doc_co_num':doc_co_num,
                                        'doc_code':doc_code,
                                        'doc_date':doc_date,
                                        'doc_value':doc_value,
                                        'id':id,
                                        'vendor_code':vendor_code,
                                        'vendor_name':vendor_name,
                                        'vendorjoint_code':vendorjoint_code,
                                        'vendorjoint_name_address':vendorjoint_name_address,
                                        'what_to_do':what_to_do,
                                        'z_co_code':z_co_code,
                                        'z_document_code':z_document_code,
                                        'modId':modId,
                                        'DocNames':DocNames ,
                                        'DocTypes':DocTypes ,
                                        'DocDescs':DocDescs ,
           })



    disb = 'CA'
    vouch = 'VO'
    insp = 'INSP'
    Disbursements = financial_Loans_lns.query('z_document_code==@disb')
    Draws = financial_Loans_lns.query('z_document_code==@vouch')
    loa = '70631447CL'
    test = Draws.query('loan_number==@loa')
    empty = ''
    test = test.query('doc_value !=@empty')
    test['doc_value'].sum()
    Inspections = financial_Loans_lns.query('z_document_code==@insp')

    financial_Loans_lns = financial_Loans_lns.query('z_document_code==@insp or z_document_code==@vouch or z_document_code==@disb')

    n = 2
    wb = Workbook()
    ws = wb.active
    ws.title = "FundControlFinancials.xlsx"
    ws['A1'] = 'loan_number'
    ws['B1'] = 'recordId'
    ws['C1'] = 'chk_dpst_account'
    ws['D1'] = 'chk_dpst_date'
    ws['E1'] = 'chk_dpst_date_cleared'
    ws['F1'] = 'chk_dpst_number'
    ws['G1'] = 'chk_dpst_number_display'
    ws['H1'] = 'chk_dpst_value'
    ws['I1'] = 'doc_co_num'
    ws['J1'] = 'doc_code'
    ws['K1'] = 'doc_date'
    ws['L1'] = 'doc_value'
    ws['M1'] = 'id'
    ws['N1'] = 'vendor_code'
    ws['O1'] = 'vendor_name'
    ws['P1'] = 'vendorjoint_code'
    ws['Q1'] = 'vendorjoint_name_address'
    ws['R1'] = 'what_to_do'
    ws['S1'] = 'z_co_code'
    ws['T1'] = 'z_document_code'
    ws['U1'] = 'modId'
    ws['V1'] = 'DocName'
    ws['W1'] = 'DocType'
    ws['X1'] = 'DocDesc'



    for index, row in financial_Loans_lns.iterrows():
        ws['A' + str(n)] = row['loan_number']
        ws['B' + str(n)] = row['recordId']
        ws['C' + str(n)] = row['chk_dpst_account']
        ws['D' + str(n)] = row['chk_dpst_date']
        ws['E' + str(n)] = row['chk_dpst_date_cleared']
        ws['F' + str(n)] = row['chk_dpst_number']
        ws['G' + str(n)] = row['chk_dpst_number_display']
        ws['H' + str(n)] = row['chk_dpst_value']
        ws['I' + str(n)] = row['doc_co_num']
        ws['J' + str(n)] = row['doc_code']
        ws['K' + str(n)] = row['doc_date']
        ws['L' + str(n)] = row['doc_value']
        ws['M' + str(n)] = row['id']
        ws['N' + str(n)] = row['vendor_code']
        ws['O' + str(n)] = row['vendor_name']
        ws['P' + str(n)] = row['vendorjoint_code']
        ws['Q' + str(n)] = row['vendorjoint_name_address']
        ws['R' + str(n)] = row['what_to_do']
        ws['S' + str(n)] = row['z_co_code']
        ws['T' + str(n)] = row['z_document_code']
        ws['U' + str(n)] = row['modId']
        ws['V' + str(n)] = row['DocNames']
        ws['W' + str(n)] = row['DocTypes']
        ws['X' + str(n)] = row['DocDescs']

        n = n + 1
    wb.save("Fund Control Financials.xlsx")



    n = 2
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "FundControlFieldData.xlsx"
    ws2['A1'] = 'loan_number'
    ws2['B1'] = 'active'
    ws2['C1'] = 'borrower'
    ws2['D1'] = 'contractor'
    ws2['E1'] = 'contract_type'
    ws2['F1'] = 'date_appraised'
    ws2['G1'] = 'date_closed'
    ws2['H1'] = 'date_end_of_draw_period'
    ws2['I1'] = 'date_inspection_last_entered'
    ws2['J1'] = 'date_loan_final'
    ws2['K1'] = 'date_matured'
    ws2['L1'] = 'department_number'
    ws2['M1'] = 'inspector'
    ws2['N1'] = 'inspection_frequency'
    ws2['O1'] = 'lender'
    ws2['P1'] = 'loan_interest_when_to_bill'
    ws2['Q1'] = 'loan_name'
    ws2['R1'] = 'loan_number_system'
    ws2['S1'] = 'loan_status'
    ws2['T1'] = 'loan_type'
    ws2['U1'] = 'project_description'
    ws2['V1'] = 'property_address'
    ws2['W1'] = 'property_address_city'
    ws2['X1'] = 'property_address_state'
    ws2['Y1'] = 'property_address_zip'
    ws2['Z1'] = 'property_type'
    ws2['AA1'] = 'revolving'
    ws2['AB1'] = 'text_notes'


    for index, row in fieldDatadf.iterrows():
        ws2['A' + str(n)] = row['loan_number']
        ws2['B' + str(n)] = row['active']
        ws2['C' + str(n)] = row['borrower']
        ws2['D' + str(n)] = row['contractor']
        ws2['E' + str(n)] = row['contract_type']
        ws2['F' + str(n)] = row['date_appraised']
        ws2['G' + str(n)] = row['date_closed']
        ws2['H' + str(n)] = row['date_end_of_draw_period']
        ws2['I' + str(n)] = row['date_inspection_last_entered']
        ws2['J' + str(n)] = row['date_loan_final']
        ws2['K' + str(n)] = row['date_matured']
        ws2['L' + str(n)] = row['department_number']
        ws2['M' + str(n)] = row['inspector']
        ws2['N' + str(n)] = row['inspection_frequency']
        ws2['O' + str(n)] = row['lender']
        ws2['P' + str(n)] = row['loan_interest_when_to_bill']
        ws2['Q' + str(n)] = row['loan_name']
        ws2['R' + str(n)] = row['loan_number_system']
        ws2['S' + str(n)] = row['loan_status']
        ws2['T' + str(n)] = row['loan_type']
        ws2['U' + str(n)] = row['project_description']
        ws2['V' + str(n)] = row['property_address']
        ws2['W' + str(n)] = row['property_address_city']
        ws2['X' + str(n)] = row['property_address_state']
        ws2['Y' + str(n)] = row['property_address_zip']
        ws2['Z' + str(n)] = row[ 'property_type']
        ws2['AA' + str(n)] = row[ 'revolving']
        ws2['AB' + str(n)] = row[ 'text_notes']
        n = n + 1

    wb2.save("Fund Control Field Data.xlsx")
    s = connectsp()
    postwb(s)
    deletewb()
    sendemailnotificationsuccess()
except Exception as inst:
    print(inst)
    sendemailnotificationfailure()

 
    
    

    
    
    
    
