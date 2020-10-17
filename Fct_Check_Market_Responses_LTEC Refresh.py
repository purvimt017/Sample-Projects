# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""
import smtplib, ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import config1

def sendemailnotificationfailure():
    sender_email = config1.username
    receiver_email = config1.username
    password = config1.password2

    message = MIMEMultipart("alternative")
    message["Subject"] = "Failed Scheduled Refresh for Fct_Check_Market_Responses_LTEC"
    message["From"] = sender_email
    message["To"] = receiver_email

    # Create the plain-text and HTML version of your message
    text = "Your scheduled refresh for Fct_Check_Market_Responses_LTEC has failed."

    # Turn these into plain/html MIMEText objects
    part1 = MIMEText(text, "plain")

    # Add HTML/plain-text parts to MIMEMultipart message
    # The email client will try to render the last part first
    message.attach(part1)

    # Create secure connection with server and send email
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(sender_email, password)
        server.sendmail(
            sender_email, receiver_email, message.as_string()
        )

try:
    import requests
    import pandas as pd
    import numpy as np
    from openpyxl import load_workbook
    import sharepy
    import os
    import config1

    def sendemailnotificationsuccess():
        sender_email = config1.username
        receiver_email = config1.username
        password = config1.password2

        message = MIMEMultipart("alternative")
        message["Subject"] = "Successful Scheduled Refresh for Fct_Check_Market_Responses_LTEC"
        message["From"] = sender_email
        message["To"] = receiver_email

        # Create the plain-text and HTML version of your message
        text = "Your scheduled refresh for Fct_Check_Market_Responses_LTEC has succeeded."

        # Turn these into plain/html MIMEText objects
        part1 = MIMEText(text, "plain")

        # Add HTML/plain-text parts to MIMEMultipart message
        # The email client will try to render the last part first
        message.attach(part1)

        # Create secure connection with server and send email
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
            server.login(sender_email, password)
            server.sendmail(
                sender_email, receiver_email, message.as_string()
            )

    def makegetrequestMT(skip):
        URL = 'https://api-us.checkmarket.com/3/surveys/83157/respondents?top=1000&skip='+ str(skip) + '&expand=Responses'
        XMaster = config1.XMaster
        XKey = config1.XKey
        HEADERS = { 'X-Master-Key':XMaster,
                   'X-Key':XKey
                   }
        req = requests.get(URL, headers=HEADERS)
        data = req.json()
        return data

    def extractdataintolists(data, dict):
        #Extract data into lists
        for i in dict:
            name = str(i)
            list = dict[str(i)]
            for d in data['Data']:
                try:
                    list.append(d[name])
                except KeyError:
                    list.append('')
    def makedf(dict, df):
        for i in dict:
            name = str(i)
            list = dict[str(i)]
            df[name] = list
        df = pd.DataFrame.from_dict(df)
        return df
    def checksizeofdata(dict):
        size = len(dict['Data'])
        return size
    def getrespondentresponsesMT(resp, df):
       lol = []
       QuestionId = []
       ResponseId = []
       ScaleValue = []
       ResponseCaption = []
       Value = []
       dffiltered = df.query('RespondentId==@resp')
       responses = dffiltered['Responses']
       for r in responses:
           if(len(r) == 0):
               QuestionId.append('')
               ResponseId.append('')
               ScaleValue.append('')
               ResponseCaption.append('')
               Value.append('')
           else:
               for i in r:
                   try:
                       QuestionId.append(i['QuestionId'])
                   except KeyError:
                       QuestionId.append('')
                   try:
                       ResponseId.append(i['ResponseId'])
                   except KeyError:
                       ResponseId.append('')
                   try:
                       ScaleValue.append(i['ScaleValue'])
                   except KeyError:
                       ScaleValue.append('')
                   try:
                       ResponseCaption.append(i['ResponseCaption'])
                   except KeyError:
                       ResponseCaption.append('')
                   try:
                       Value.append(str(i['Value']))
                   except KeyError:
                       Value.append('')
       lol.append(QuestionId)
       lol.append(ResponseId)
       lol.append(ScaleValue)
       lol.append(ResponseCaption)
       lol.append(Value)
       return lol
    def insert_row(idx, df, df_insert):
        dfA = df.iloc[:idx, ]
        dfB = df.iloc[idx:, ]
        df = dfA.append(df_insert).append(dfB).reset_index(drop = True)
        return df
    def addvaluestolist(list, Id):
        for x in list:
            Id.append(x)
    def createcolumnsMT(QuestionId,ResponseId,ScaleValue,ResponseCaption,Value,df):
        df['QuestionId'] = QuestionId
        df['ResponseId'] = ResponseId
        df['ScaleValue'] = ScaleValue
        df['ResponseCaption'] = ResponseCaption
        df['Value'] = Value
        df = df.drop(['Responses'], axis = 1)
        return df
    def mastertablebeforeresp(skip):
        responsesdf = {}
        df = pd.DataFrame({})
        while True:
            data = makegetrequestMT(skip)
            size = checksizeofdata(data)
            extractdataintolists(data, dict)
            makedf(dict, responsesdf)
            if(size != 1000):
                break
            skip += 1000
        df = pd.DataFrame.from_dict(responsesdf)
        return df
    #build a larger DF to insert instead of one at a time
    def extractresponsesMT(QuestionId, ResponseId, ScaleValue, ResponseCaption, Value, df):
        index1 = 0
        for index, row in df.iterrows():
            Responseslist = getrespondentresponsesMT(row['RespondentId'], df)
            Questionlist = Responseslist[0]
            Responselist = Responseslist[1]
            ScaleValuelist = Responseslist[2]
            ResponseCaptionlist = Responseslist[3]
            Valuelist = Responseslist[4]
            addvaluestolist(Questionlist, QuestionId)
            addvaluestolist(Responselist, ResponseId)
            addvaluestolist(ScaleValuelist, ScaleValue)
            addvaluestolist(ResponseCaptionlist, ResponseCaption)
            addvaluestolist(Valuelist, Value)
            length1 = len(Questionlist)
            if(length1 > 0):
                length2 = length1 - 1
            else:
                length2 = 0
            df1 = pd.DataFrame({}) #create new empty dataframe
            for c in df.columns:
                df1[c] = np.zeros(length2) #for column in df2, df1(column) = an array the lenth of lenth2
            df2 = pd.DataFrame(row).transpose() #transpose the row from df2 and create a new df with it
            for n in range(length2):
                df1.loc[n,:] = df2.iloc[0] #for every number in range of length2, df.loc[0,:] = row ath location 0
            df = insert_row(index1, df, df1.iloc[:length2, ])
            if(length1 == 0):
                index1 += 1
            else:
                index1 += length1
            print(index1)
        return df
#    def determineskip(SurveyNumber):
#        surveynumber = str(SurveyNumber)
#       skip = surveynumber[:2]
#        skip = int(skip) * 1
#        if SurveyNumber % 1000 == 0:
#            skip += 1000
#        else :
#            skip = skip
#        return skip
    def returnlatestdata(df):
        df2 = pd.DataFrame({})
        for c in df.columns:
            df2[c] = np.zeros(0)
        for index, row in df.iterrows():
            if row['RespondentId'] > SurveyNumber:
                df3 = pd.DataFrame(row).transpose()
                df2 = df2.append(df3)
        return df2
    def getfile(s, filename, url, *args, **kwargs):
            """Stream download of specified URL and output to file"""
            # Extract file name from request URL if not provided as keyword argument
            #filename = kwargs.pop("filename", re.search(r"[^/]+$", url).group(0))
            kwargs["stream"] = True
            # Request file in stream mode
            response = s.get(url, *args, **kwargs)
            # Save to output file
            if response.status_code == requests.codes.ok:
                with open(filename, "wb") as file:
                    for chunk in response:
                        file.write(chunk)
            file = pd.read_excel(filename)
            return file
    def downloadlatestwb():
        s = sharepy.connect("https://ltsystems.sharepoint.com", username = config1.username, password = config1.password)
        file = getfile(s,'Fct_Check_Market_Responses_LTEC.xlsx',
                       'https://ltsystems.sharepoint.com/sites/PowerBIAdmin45/Shared%20Documents/Source_Data_Files/Fct_Check_Market_Responses_LTEC.xlsx')
        return s, file
    def getNandSurveyNumber(file):
        rn = file['Survey Number']
        n = rn.count() + 2
        SurveyNumber = rn.iloc[-1]
        return n, SurveyNumber
    def editwb(n):
        wb = load_workbook('Fct_Check_Market_Responses_LTEC.xlsx')
        ws = wb['MasterResponses']
        for index, row in df2.iterrows():
            ws['A' + str(n)] = row['RespondentId']
            ws['B' + str(n)] = row['RespondentStatusId']
            ws['C' + str(n)] = row['ResponseDate']
            ws['D' + str(n)] = row['CompletionTime']
            ws['E' + str(n)] = row['RespondentReportUrl']
            ws['F' + str(n)] = row['ReportUrl']
            ws['G' + str(n)] = row['QuestionId']
            ws['H' + str(n)] = row['ResponseId']
            ws['I' + str(n)] = row['Value']
            ws['J' + str(n)] = row['RespondentHash']
            ws['K' + str(n)] = row['PanelistStatusId']
            ws['L' + str(n)] = row['DateAdded']
            ws['M' + str(n)] = row['DateInvited']
            ws['N' + str(n)] = row['DateSawMail']
            ws['O' + str(n)] = row['DateClickedThrough']
            ws['P' + str(n)] = row['DateResponded']
            ws['Q' + str(n)] = row['DateToBeInvited']
            ws['R' + str(n)] = row['DateLastModified']
            ws['S' + str(n)] = row['ContactId']
            ws['T' + str(n)] = row['FirstName']
            ws['U' + str(n)] = row['LastName']
            ws['V' + str(n)] = row['Email']

            n = n + 1
        wb.save("Fct_Check_Market_Responses_LTEC.xlsx")
    def postwb(s):
        with open('Fct_Check_Market_Responses_LTEC.xlsx', 'rb') as file_input:
            try:
                response = s.post(
                    url='https://ltsystems.sharepoint.com' + "/sites/" + "PowerBIAdmin45"
                    + "/_api/web/GetFolderByServerRelativeUrl('" + 'Shared%20Documents/Source_Data_Files' + "')/Files/add(url='"
                    + 'Fct_Check_Market_Responses_LTEC.xlsx' + "',overwrite=true)",
                    data=file_input)
                status_code = response.status_code
                print("{}: Upload Successful".format(status_code))
            except Exception as err:
                print("Some error occurred: " + str(err))
    def deletewb():
        if os.path.exists("Fct_Check_Market_Responses_LTEC.xlsx"):
          os.remove("Fct_Check_Market_Responses_LTEC.xlsx")
        else:
          print("The file does not exist")


    #Initialize Lists for Columns
    RespondentId = []
    LanguageCode = []
    RespondentStatusId = []
    ResponseDate = []
    CompletionTime = []
    DistributionMethodId = []
    BrowserId = []
    OsId = []
    IpAddress = []
    RespondentReportUrl = []
    ReportUrl = []
    Responses = []
    RespondentHash = []
    PanelistStatusId = []
    DateAdded = []
    DateInvited = []
    DateSawMail = []
    DateClickedThrough = []
    DateResponded = []
    DateLastModified = []
    DateToBeInvited = []
    ContactId = []
    FirstName = []
    LastName = []
    Email = []
    Password = []
    IsMobile = []
    #Initialize lists for responses
    QuestionId = []
    ResponseId = []
    ScaleValue = []
    ResponseCaption = []
    Value = []

    #Create Dictionary for Columns
    dict = {'RespondentId':RespondentId,
            'LanguageCode':LanguageCode,
            'RespondentStatusId':RespondentStatusId,
            'ResponseDate':ResponseDate,
            'CompletionTime':CompletionTime,
            'DistributionMethodId':DistributionMethodId,
            'BrowserId':BrowserId,
            'OsId':OsId,
            'IpAddress':IpAddress,
            'RespondentReportUrl':RespondentReportUrl,
            'ReportUrl':ReportUrl,
            'Responses':Responses,
            'RespondentHash':RespondentHash,
            'PanelistStatusId':PanelistStatusId,
            'DateAdded':DateAdded,
            'DateInvited':DateInvited,
            'DateSawMail':DateSawMail,
            'DateClickedThrough':DateClickedThrough,
            'DateResponded':DateResponded,
            'DateLastModified':DateLastModified,
            'DateToBeInvited':DateToBeInvited,
            'ContactId':ContactId,
            'FirstName':FirstName,
            'LastName':LastName,
            'Email':Email,
            'IsMobile':IsMobile
            }
    #download Sharepoint File
    s, file = downloadlatestwb()
    #Get the last Survey Number and last n
    n, SurveyNumber = getNandSurveyNumber(file)
    #Determine Skip
    skip = determineskip(SurveyNumber)
    #Get latest Data
    df = mastertablebeforeresp(skip)
    #Build df with latestdata not on excel
    df2 = returnlatestdata(df)
    #extract the responses
    df2 = extractresponsesMT(QuestionId, ResponseId, ScaleValue, ResponseCaption, Value, df2)
    #Create new columns
    df2 = createcolumnsMT(QuestionId,ResponseId,ScaleValue,ResponseCaption,Value, df2)
    #Edit Workbook
    editwb(n)
    #Post Workbook
    postwb(s)
    #Delete workbook in local filepath
    deletewb()
    sendemailnotificationsuccess()
except Exception as inst:
    print(inst)
    sendemailnotificationfailure()



